"""Parse Timing App Excel export and yield rows with Day, Start, End, Project, Title.

Expected headers match Timing's export and the `export_timing_report` CLI output:
Start Date, End Date, Project, Title.
"""
from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path

import openpyxl

# Excel serial date epoch (Lotus 1-2-3 / Excel compatibility)
_EXCEL_EPOCH = datetime(1899, 12, 30)


def _to_datetime(value) -> datetime | None:
    """Convert Excel serial (float) or existing datetime to datetime."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    days = int(serial)
    frac = serial - days
    base = _EXCEL_EPOCH + timedelta(days=days)
    if frac <= 0:
        return base
    seconds = round(frac * 24 * 3600)
    return base + timedelta(seconds=seconds)


def _format_day(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%d")


def _format_time(dt: datetime) -> str:
    """Format as HH:MM."""
    return dt.strftime("%H:%M")


def parse_timing_export(path: str | Path) -> list[dict]:
    """
    Read the first sheet of a Timing App XLSX export and yield one dict per data row
    with keys: day (YYYY-MM-DD), start (H:MM), end (H:MM), project, title.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Timing export not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return []

    # Row 1 = headers; find column indices by name
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        wb.close()
        return []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    col = {h: i for i, h in enumerate(headers)}

    required = ("Start Date", "End Date", "Project", "Title")
    for r in required:
        if r not in col:
            wb.close()
            raise ValueError(f"Timing export missing column: {r}")

    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        # Pad so we can index by col[name]
        while len(row) < len(headers):
            row.append(None)
        if not row:
            continue

        def _val(key: str, default=None):
            i = col.get(key)
            if i is None or i >= len(row):
                return default
            v = row[i]
            return v if v is not None else default

        start_dt = _to_datetime(_val("Start Date"))
        if start_dt is None:
            continue

        end_dt = _to_datetime(_val("End Date")) or start_dt
        project = _val("Project") or ""
        title = _val("Title") or ""

        out.append({
            "day": _format_day(start_dt),
            "start": _format_time(start_dt),
            "end": _format_time(end_dt),
            "project": str(project).strip(),
            "title": str(title).strip(),
        })

    wb.close()
    return out
