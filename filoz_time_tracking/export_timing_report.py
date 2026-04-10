"""CLI: export Timing report to XLSX for FilOz project and date range.

Read-only with respect to Timing data: this module does not call the Timing Web API and does not
create, update, or delete time entries or projects. It only runs TimingHelper's ``save report`` to
write a local workbook. The JXA ``helper.delete(reportSettings/exportSettings)`` calls release
in-memory scripting objects (per Timing's samples), not user records.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import shlex
import subprocess
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Iterable

import openpyxl


class UserInputError(ValueError):
    """Raised for invalid CLI arguments or unsupported user input."""


class RuntimeExportError(RuntimeError):
    """Raised for Timing export runtime failures."""


@dataclass(frozen=True)
class DateRange:
    start: date
    end: date


_INVOICE_RE = re.compile(r"^(\d{4})-(\d{1,2})$")


def _parse_date(value: str, flag_name: str) -> date:
    try:
        return date.fromisoformat(value)
    except ValueError as exc:
        raise UserInputError(f"{flag_name} must be YYYY-MM-DD (got: {value!r})") from exc


def _parse_invoice_period(value: str) -> DateRange:
    match = _INVOICE_RE.match(value.strip())
    if not match:
        raise UserInputError("--invoice must be YYYY-M or YYYY-MM (e.g. 2026-3)")

    year = int(match.group(1))
    month = int(match.group(2))
    if month < 1 or month > 12:
        raise UserInputError("--invoice month must be between 1 and 12")

    if month == 1:
        start = date(year - 1, 12, 10)
    else:
        start = date(year, month - 1, 10)
    end = date(year, month, 9)
    return DateRange(start=start, end=end)


def _resolve_range(args: argparse.Namespace) -> DateRange:
    has_explicit = bool(args.start or args.end)
    has_invoice = bool(args.invoice)

    if has_invoice and has_explicit:
        raise UserInputError("Use either --invoice or --start/--end, not both.")

    if has_invoice:
        return _parse_invoice_period(args.invoice)

    if not args.start and not args.end:
        raise UserInputError("Either --invoice or both --start and --end is required.")
    if not args.start or not args.end:
        raise UserInputError("--start and --end must be provided together.")

    start = _parse_date(args.start, "--start")
    end = _parse_date(args.end, "--end")
    if end < start:
        raise UserInputError("--end must be on or after --start.")
    return DateRange(start=start, end=end)


def _default_output_path(project: str, export_range: DateRange) -> Path:
    project_slug = re.sub(r"[^A-Za-z0-9_-]+", "-", project.strip()).strip("-") or "timing"
    filename = f"{project_slug}-{export_range.start.isoformat()}_{export_range.end.isoformat()}.xlsx"
    return Path.cwd() / filename


def _build_jxa_source(start_iso: str, end_iso: str, output_path: str, project_name: str) -> str:
    """TimingHelper JXA: values embedded via JSON to avoid argv and quoting issues."""
    out_path = json.dumps(output_path)
    proj = json.dumps(project_name)
    start_lit = json.dumps(start_iso)
    end_lit = json.dumps(end_iso)
    return f"""
var helper = Application("TimingHelper");
if (!helper.advancedScriptingSupportAvailable()) {{
  throw "This script requires a Timing Connect subscription. Please contact support via https://timingapp.com/contact to upgrade.";
}}
var startDate = new Date({start_lit} + "T00:00:00");
var endDate = new Date({end_lit} + "T23:59:59");
var outputPathStr = {out_path};
var projectFilter = {proj};

var reportSettings = helper.ReportSettings().make();
var exportSettings = helper.ExportSettings().make();

reportSettings.firstGroupingMode = "raw";
reportSettings.timeEntriesIncluded = true;
reportSettings.appUsageIncluded = false;

exportSettings.fileFormat = "Excel";
exportSettings.durationFormat = "hhmmss";
exportSettings.shortEntriesIncluded = true;

var saveArgs = {{
  withReportSettings: reportSettings,
  exportSettings: exportSettings,
  between: startDate,
  and: endDate,
  to: Path(outputPathStr)
}};
if (projectFilter.length > 0) {{
  saveArgs.forProjects = helper.projects.whose({{ name: projectFilter }});
  saveArgs.withSubprojectsIncluded = true;
}}
helper.saveReport(saveArgs);

// Release report/export setting handles only (avoids AppleScript memory leak); does not delete data.
helper.delete(reportSettings);
helper.delete(exportSettings);
"""


def _applescript_escape_double_quoted(s: str) -> str:
    """Escape for use inside an AppleScript string literal delimited by double quotes."""
    return s.replace("\\", "\\\\").replace('"', '\\"')


def _run_timing_export(
    output_path: Path,
    export_range: DateRange,
    project_name: str,
    timeout_seconds: int,
) -> None:
    """Run JXA via osascript, wrapped in AppleScript `with timeout` to avoid -1712 on large exports."""
    js_source = _build_jxa_source(
        export_range.start.isoformat(),
        export_range.end.isoformat(),
        str(output_path.resolve()),
        project_name.strip(),
    )
    fd, js_path = tempfile.mkstemp(suffix=".js", prefix="filoz-timing-export-")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(js_source)
        inner = " ".join(
            shlex.quote(x) for x in ["/usr/bin/osascript", "-l", "JavaScript", js_path]
        )
        # AppleScript requires a double-quoted string for `do shell script`; do not use shlex.quote
        # here (single quotes are invalid AppleScript string delimiters and yield -2741).
        body = _applescript_escape_double_quoted(inner)
        applescript = (
            f"with timeout of {timeout_seconds} seconds\n"
            f'\tdo shell script "{body}"\n'
            "end timeout"
        )
        result = subprocess.run(
            ["osascript", "-e", applescript],
            capture_output=True,
            text=True,
        )
    finally:
        try:
            os.unlink(js_path)
        except OSError:
            pass

    if result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()
        if "Timing Connect subscription" in detail:
            raise UserInputError(detail)
        if "-1712" in detail or "timed out" in detail.lower():
            raise RuntimeExportError(
                f"Timing export timed out (Apple Event -1712). Try: Timing app running and responsive; "
                f"a smaller date range; `--timeout` (current {timeout_seconds}s); `--project` to limit scope. Raw: {detail}"
            )
        raise RuntimeExportError(f"Timing export failed via osascript: {detail or 'unknown error'}")


def _to_datetime(value: object) -> datetime | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.fromisoformat(str(value))
    except ValueError:
        pass
    try:
        serial = float(value)
    except (TypeError, ValueError):
        return None
    epoch = datetime(1899, 12, 30)
    days = int(serial)
    frac = serial - days
    return epoch + timedelta(days=days) + timedelta(seconds=round(frac * 24 * 3600))


def _is_project_match(project_value: str, root_project: str) -> bool:
    if not project_value:
        return False
    project_value = project_value.strip()
    root = root_project.strip()
    if project_value == root:
        return True
    prefixes = (f"{root} ▸", f"{root}/", f"{root}:")
    return any(project_value.startswith(prefix) for prefix in prefixes)


def _iter_filtered_rows(
    rows: Iterable[tuple],
    col: dict[str, int],
    root_project: str,
    export_range: DateRange,
) -> list[tuple]:
    out: list[tuple] = []
    for row in rows:
        row_data = list(row)
        while len(row_data) <= max(col.values()):
            row_data.append(None)
        start_dt = _to_datetime(row_data[col["Start Date"]])
        end_dt = _to_datetime(row_data[col["End Date"]])
        project = str(row_data[col["Project"]] or "")
        if not start_dt or not end_dt:
            continue
        if not _is_project_match(project, root_project):
            continue
        if start_dt.date() > export_range.end or end_dt.date() < export_range.start:
            continue
        out.append(tuple(row_data))
    return out


def _filter_export_file(output_path: Path, root_project: str, export_range: DateRange) -> int:
    wb = openpyxl.load_workbook(output_path, data_only=False)
    ws = wb.active
    if ws is None:
        wb.close()
        raise RuntimeExportError("Export workbook has no active sheet.")

    header = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), [])]
    col = {name: i for i, name in enumerate(header)}
    required = ("Start Date", "End Date", "Project", "Title")
    for req in required:
        if req not in col:
            wb.close()
            raise RuntimeExportError(f"Timing export missing expected column: {req}")

    rows = _iter_filtered_rows(
        ws.iter_rows(min_row=2, values_only=True),
        col=col,
        root_project=root_project,
        export_range=export_range,
    )

    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.append(header)
    for row in rows:
        out_ws.append(row)
    out_wb.save(output_path)
    wb.close()
    out_wb.close()
    return len(rows)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Export Timing report to XLSX for a date range or invoice period. "
            "Uses local TimingHelper automation only (no Timing Web API; no mutations to time data)."
        ),
    )
    parser.add_argument(
        "-o",
        "--output",
        type=str,
        default=None,
        help="Output XLSX path. Default: ./<project>-<start>_<end>.xlsx",
    )
    parser.add_argument("--start", type=str, default=None, help="Inclusive start date (YYYY-MM-DD)")
    parser.add_argument("--end", type=str, default=None, help="Inclusive end date (YYYY-MM-DD)")
    parser.add_argument(
        "--invoice",
        type=str,
        default=None,
        help="Invoice shorthand YYYY-M or YYYY-MM (period: previous month 10th to month 9th).",
    )
    parser.add_argument(
        "--project",
        type=str,
        default="FilOz",
        help="Root project name to include, with all subprojects (default: FilOz).",
    )
    parser.add_argument(
        "--timeout",
        type=int,
        default=900,
        metavar="SECONDS",
        help="Apple Event timeout for Timing export (default: 900). Increase if you see -1712.",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    try:
        export_range = _resolve_range(args)
        output_path = Path(args.output).expanduser() if args.output else _default_output_path(args.project, export_range)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        if args.timeout < 60:
            raise UserInputError("--timeout must be at least 60 seconds.")
        _run_timing_export(
            output_path=output_path,
            export_range=export_range,
            project_name=args.project,
            timeout_seconds=args.timeout,
        )
        row_count = _filter_export_file(output_path=output_path, root_project=args.project, export_range=export_range)

        print(f"Export complete: {output_path}")
        print(f"Project root: {args.project}")
        print(f"Effective range: {export_range.start.isoformat()} to {export_range.end.isoformat()} (inclusive)")
        if row_count == 0:
            print("No matching entries found for the selected project/range. Wrote headers-only workbook.")
        else:
            print(f"Rows retained: {row_count}")
        return 0
    except UserInputError as exc:
        print(exc, file=sys.stderr)
        return 1
    except RuntimeExportError as exc:
        print(exc, file=sys.stderr)
        return 2
    except Exception as exc:  # pragma: no cover
        print(f"Unexpected error: {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
